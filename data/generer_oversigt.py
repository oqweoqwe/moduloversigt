import openpyxl
from datetime import datetime
import time
from os.path import exists
from os import remove
from sys import exit

# flerklassehold: 2cf Mu, 2fi MA, 3c4i EN 1
# valgfag

illegal = ["1g","blå","grøn","rød","gul","gf","dho","ff","sro","ssb","øh","hørelære","sammenspil","kor","klassisk","rytmisk","mgk","projekttime","kt","eksamen"]

log_path = "latest.txt"

holdrapport = "data/Samlet holdrapport.xlsx"
modulfordeling = "data/Samlet modulfordeling.xlsx"

classes = ["2a","2b","2c","2d","2e","2f","2h","2i","3a","3b","3c","3d","3e","3f","3h","3i","4i"]
subjects = {"da":"dansk","dr":"drama", "bi":"biologi", "en":"engelsk","hi":"historie","id":"idræt","ma":"matematik","mu":"musik","sa":"samfundsfag",
			"fy":"fysik", "bt":"biotek","ke":"kemi","ap":"ap","apla":"apla","ps":"psykologi","la":"latin","bk":"billedkunst","fi":"filosofi",
			"ty":"tysk","fr":"fransk","sp":"spansk","ng":"naturgeografi","nv":"nv","ol":"oldtidskundskab","re":"religion"}

class Main:

	def __init__(self):

		self.teams = []
		self.result_path = "Oversigt.xlsx"
		self.result = openpyxl.Workbook()
		self.klasser = []
		self.valgfag = []
		self.flerklassehold = []

		self.editing_sheet = None

		# if latest.txt exists, delete it
		if exists(log_path):
			self.log("Found old log, deleting it")
			remove(log_path)

		# load files into memory
		self.load_files()

		# get all teams from team report, filtering out unimportant ones
		self.load_teams()

		# load stamklasser
		self.load_klasser()

		# load the correct amount of modules for each team from the plan
		self.load_correct_amounts()

		# write
		self.write()

	def write(self):

		self.error = []

		self.log(f"Started writing to {self.result_path}")

		# write stamklasseundervisning
		for klasse in classes:
			self.setup_editing_sheet(klasse)
			for team in self.get_klasse_by_name(klasse).teams:
				
				self.write_team(team)
		
		# write flerklassehold
		for team in self.flerklassehold:
			
			# hvis sheet eksisterer
			if "Andre" in self.result.sheetnames:
				self.editing_sheet = self.result["Andre"]
				self.write_team(team)
			else:
				self.setup_editing_sheet("Andre")
				self.write_team(team)

		# write valgfag
		for team in self.valgfag:
			
			# hvis sheet eksisterer
			if "Valgfag" in self.result.sheetnames:
				self.editing_sheet = self.result["Valgfag"]
				self.write_team(team)
			else:
				self.setup_editing_sheet("Valgfag")
				self.write_team(team)

		self.style_cells()

		self.result.remove(self.result["Sheet"])

		self.result.save(self.result_path)

		if len(self.error) > 0:
			self.log(f"An error occured while parsing the following teams ({len(self.error)}/{len(self.teams)}): {self.error}")

		self.log(f"Results saved at {self.result_path}")

	def style_cells(self):

		for sheet_name in self.result.sheetnames:
			sheet = self.result[sheet_name]
			for cell in sheet['D']:
				if cell.row == 1:
					for cell in sheet[cell.row]:
						if cell:
							cell.style = "Headline 1"
					continue
				if cell.value == 0:
					for cell in sheet[cell.row]:
						if cell:
							cell.style = "Good"
					continue
				elif cell.value < 4:
					for cell in sheet[cell.row]:
						if cell:
							cell.style = "Neutral"
					continue
				else:
					for cell in sheet[cell.row]:
						if cell:
							cell.style = "Bad"
					continue

	def write_team(self, team):
		self.editing_sheet[f"A{self.index}"].value = team.name
		self.editing_sheet[f"B{self.index}"].value = team.total
		planned_amount = team.planned_amount
		self.editing_sheet[f"C{self.index}"].value = planned_amount
		self.editing_sheet[f"D{self.index}"].value = team.total - planned_amount

		if planned_amount == -1:
			self.error.append(team.name)
		self.index = self.index+1

	def setup_editing_sheet(self, name):
		self.editing_sheet = self.result.create_sheet(name)
		self.editing_sheet["A1"].value = "Hold"
		self.editing_sheet["B1"].value = "Skema"
		self.editing_sheet["C1"].value = "Norm"
		self.editing_sheet["D1"].value = "Afvigelse"
		self.index = 2

	def get_klasse_by_name(self, name):
		for klasse in self.klasser:
			if klasse.name == name:
				return klasse
		return None

	def load_klasser(self):
		for klasse in classes:
			self.klasser.append(Klasse(klasse, self.teams))

	def log(self, msg):
		now = datetime.now().strftime("%H:%M:%S")
		txt = f"[{now}] {msg}"
		print(txt)
		with open(log_path, 'a') as f:
			f.write("".join(txt))
			f.write("\n")

	def load_files(self):
		
		if exists(holdrapport):
			before = time.time()
			self.log(f"Loading {holdrapport}...")
			self.team_report = openpyxl.load_workbook(holdrapport)
			self.team_report_sheet = self.team_report.active
			self.log(f"Load took {round(time.time()-before,1)} seconds")
		else:
			self.log(f"Couldn't find {holdrapport}, exiting")
			exit()

		if exists(modulfordeling):
			before = time.time()
			self.log(f"Loading {modulfordeling}...")
			self.module_distribution = openpyxl.load_workbook(modulfordeling)
			self.log(f"Load took {round(time.time()-before,1)} seconds")
		else:
			self.log(f"Couldn't find {modulfordeling}, exiting")
			exit()

		self.log(f"Using the following team report: {self.team_report_sheet['A1'].value}")

	def load_teams(self):
		
		for i in range(4,self.team_report_sheet.max_row+1-5):
			name = self.team_report_sheet[f'A{i}'].value
			total = self.team_report_sheet[f'G{i}'].value

			if self.is_name_legal(name):
				self.teams.append(Team(name,total))

	def get_level(self, team_name):
		subject_id = team_name.split(" ",2)[1]

		if subject_id.islower():
			return "C"
		if subject_id.isupper():
			return "A"
		return "B"

	def set_planned_amount(self, column, occurences, cell, sheet, subject, team):

		# hvis faget står flere gange, vælg det rigtige niveau
		if occurences > 2:
			for cell in sheet['D']:
				if subject in str(cell.value).lower():
					if f" {team.level}" in str(cell.value) or f"/{team.level}" in str(cell.value):
						if team.planned_amount == -1:
							team.planned_amount = sheet[f'{column}{cell.row}'].value

		elif team.planned_amount == -1:
			team.planned_amount = sheet[f'{column}{cell.row}'].value

	# also loads subject levels
	def load_correct_amounts(self):
		for team in self.teams:
			if team.name.split(' ',1)[0] in classes:
				# team er stamklasseundervisning

				class_id = self.get_class_id(team.name)
				subject = self.get_subject_name(team.name)
				sheet = self.get_sheet_by_class_id(class_id)

				team.level = self.get_level(team.name)

				# håndter 2. fremmedsprog i stamklasser, eks. 2a Ty
				if subject == "tysk" or subject == "fransk":
					subject = "fremmedsprog"

				occurences = 0

				# count occurences
				for cell in sheet['D']:
					if subject in str(cell.value).lower():
						occurences = occurences+1

				for cell in sheet['D']:
					if subject in str(cell.value).lower():

						if not sheet[f"E{cell.row}"].value == "Undervisning":
								continue

						if self.get_year_from_class_id(class_id) == 1:
							self.set_planned_amount('F', occurences, cell, sheet, subject, team)
							continue
						if self.get_year_from_class_id(class_id) == 2:
							self.set_planned_amount('H', occurences, cell, sheet, subject, team)
							continue
						if self.get_year_from_class_id(class_id) == 3:
							self.set_planned_amount('J', occurences, cell, sheet, subject, team)
							continue
						if self.get_year_from_class_id(class_id) == 4:
							self.set_planned_amount('L', occurences, cell, sheet, subject, team)
							continue
			else:
				# team er ikke stamklasseundervisning
				if not 'g' in team.name.split(" ",1)[0]:
					# flerklassehold, eks. 2cf Mu
					self.flerklassehold.append(team)
				else:
					# valgfag
					self.valgfag.append(team)

	def get_subject_name(self, team_name):
		return subjects.get(team_name.lower().split(" ",2)[1])

	def get_year_from_class_id(self, class_id):

		if int(class_id[0:4]) == datetime.today().year:
			return 1
		if int(class_id[0:4]) == datetime.today().year-1:
			return 2
		if int(class_id[0:4]) == datetime.today().year-2:
			return 3
		if int(class_id[0:4]) == datetime.today().year-3:
			return 4

	def get_sheet_by_class_id(self, class_id):

		sheets = self.module_distribution.sheetnames
		for sheet in sheets:
			if class_id in sheet:
				return self.module_distribution[sheet]

	def get_class_id(self, name):

		if '4' in name:
			return str(datetime.today().year-3)+name[1].lower()
		if '3' in name:
			return str(datetime.today().year-2)+name[1].lower()
		if '2' in name:
			return str(datetime.today().year-1)+name[1].lower()
		if '1' in name:
			return str(datetime.today().year-0)+name[1].lower()

	def is_name_legal(self, name):

		if name[0] == "1":
			return False

		for word in illegal:
			if word in name.lower():
				return False
		return True

class Team:

	def __init__(self, name, total):

		self.name = name
		self.total = total
		self.planned_amount = -1
		self.level = None

class Klasse:

	def __init__(self, name, all_teams):

		self.name = name
		self.teams = []
		for team in all_teams:
			if team.name.split(' ',1)[0] in classes and team.name.split(" ",1)[0] == name:
				self.teams.append(team)

if __name__ == '__main__':
	main = Main()